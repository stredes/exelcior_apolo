from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from pathlib import Path

def save_pretty_excel(df, output_filename):
    """Guarda el DataFrame editado con formato personalizado."""
    wb = Workbook()
    ws = wb.active

    # Estilos
    header_fill = PatternFill("solid", fgColor="00B0F0")  # Azul claro
    total_fill = PatternFill("solid", fgColor="00B0F0")
    bold_font_white = Font(bold=True, color="FFFFFF")
    bold_font_black = Font(bold=True, color="000000")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Título con fecha
    today = datetime.now().strftime("%A, %d de %B de %Y").capitalize()
    ws.merge_cells('A1:F1')
    ws['A1'] = f"AMILAB-FEDEX      {today}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Encabezados
    headers = list(df.columns)
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font_white
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # TOTAL BULTOS
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=5).value = "TOTAL BULTOS"
    ws.cell(row=total_row, column=5).font = bold_font_black
    ws.cell(row=total_row, column=5).fill = total_fill
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=6).value = df['BULTOS'].sum()
    ws.cell(row=total_row, column=6).font = bold_font_black
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="center")

    # Ajustar ancho columnas (evitar merged cells)
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        ws.column_dimensions[col_letter].width = max_length + 2

    # Bordes y alineación para todos los datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Guardar
    output_path = Path.cwd() / output_filename
    wb.save(output_path)
    return output_path
