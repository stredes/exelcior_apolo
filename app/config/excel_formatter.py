from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import pandas as pd

# --- Helpers para título en español (sin depender de locale) ---
_DIAS = ["lunes","martes","miércoles","jueves","viernes","sábado","domingo"]
_MESES = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]

def _fecha_es(dt: datetime) -> str:
    d = dt.weekday()  # 0=lunes
    return f"{_DIAS[d].capitalize()}, {dt.day:02d} de {_MESES[dt.month-1]} de {dt.year}"

def save_pretty_excel(
    df: pd.DataFrame,
    output_filename: str = "fedex_listado.xlsx",
    output_dir: str | Path | None = None,
    titulo_prefix: str = "AMILAB-FEDEX"
) -> Path:
    """
    Guarda el DataFrame con formato “listado profesional” y evita errores de MergedCell.
    - Encabezados con color, título con fecha en español.
    - Anchos automáticos (desde fila 2 para no tocar celdas fusionadas).
    - Congela encabezados, auto-filtro y bordes en datos y totales.
    - Permite elegir carpeta de salida.

    Parameters
    ----------
    df : pd.DataFrame
        Debe contener la columna 'BULTOS' (int).
    output_filename : str
        Nombre del archivo a generar.
    output_dir : str | Path | None
        Carpeta de salida; si None usa el cwd.
    titulo_prefix : str
        Texto a la izquierda del título antes de la fecha.
    """
    if df is None or df.empty:
        raise ValueError("El DataFrame está vacío; no hay nada para exportar.")
    if "BULTOS" not in df.columns:
        raise KeyError("La columna requerida 'BULTOS' no existe en el DataFrame.")

    # Normaliza BULTOS a int (evita floats al sumar)
    df = df.copy()
    df["BULTOS"] = pd.to_numeric(df["BULTOS"], errors="coerce").fillna(0).astype(int)

    wb = Workbook()
    ws = wb.active
    ws.title = "FedEx"

    # Estilo base
    header_fill = PatternFill("solid", fgColor="00B0F0")  # azul claro
    total_fill = PatternFill("solid", fgColor="00B0F0")
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True, color="000000")
    base_font = Font(name="Segoe UI", size=10)
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Formato por defecto de filas (altura agradable)
    ws.sheet_format = SheetFormatProperties(defaultRowHeight=18.0)

    # --- Título (fila 1) con merge sin tocar la fila 2 en cálculos de ancho ---
    n_cols = max(1, len(df.columns))
    fecha_txt = _fecha_es(datetime.now())
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1).value = f"{titulo_prefix}      {fecha_txt}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # --- Encabezados (fila 2) ---
    headers = list(df.columns)
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx)
        c.fill = header_fill
        c.font = bold_white
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin

    # --- Datos (desde fila 3) ---
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Bordes/centrado datos
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(3, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = base_font
            cell.border = border_thin
            # Centrado en general; si prefieres izq excepto BULTOS, cambia aquí.
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Total BULTOS (fila siguiente) ---
    total_row = ws.max_row + 1
    # Etiqueta en penúltima columna y valor en última (siempre última=columna BULTOS)
    label_col = max(1, max_col - 1)
    value_col = max_col

    ws.cell(row=total_row, column=label_col, value="TOTAL BULTOS")
    ws.cell(row=total_row, column=value_col, value=int(df["BULTOS"].sum()))

    for c in (label_col, value_col):
        cell = ws.cell(row=total_row, column=c)
        cell.font = bold_black
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    # --- Anchos automáticos (desde fila 2 para evitar merged de la fila 1) ---
    #     Esto elimina el riesgo de 'MergedCell' object no tiene 'column_letter'.
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        # Recorremos cabecera y datos (fila 2..total_row)
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

    # --- Calidad de vida ---
    ws.freeze_panes = "A3"                 # Congela título+encabezados
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"  # Filtro en encabezados
    # Márgenes de impresión cómodos
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    # --- Guardar ---
    out_dir = Path(output_dir) if output_dir else Path.cwd()
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / output_filename
    wb.save(output_path)
    return output_path
