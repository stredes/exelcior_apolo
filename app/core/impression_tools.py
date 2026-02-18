# app/core/impression_tools.py

import os
import platform
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Optional

import pandas as pd
from app.utils.utils import autoajustar_columnas
from app.core.logger_eventos import log_evento


def _windows_printer_names() -> list[str]:
    if platform.system() != "Windows":
        return []
    try:
        import win32print  # type: ignore
        flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        names = []
        for item in win32print.EnumPrinters(flags):
            try:
                n = str(item[2]).strip()
            except Exception:
                continue
            if n:
                names.append(n)
        return list(dict.fromkeys(names))
    except Exception:
        return []


def _resolve_windows_printer_name(alias: str) -> str:
    base = (alias or "").strip()
    if not base or platform.system() != "Windows":
        return base
    names = _windows_printer_names()
    if not names:
        return base
    low = base.lower()
    for n in names:
        if n.lower() == low:
            return n
    for n in names:
        if low in n.lower() or n.lower() in low:
            return n
    return base


def _excel_active_printer_candidates(printer_name: str) -> list[str]:
    base = _resolve_windows_printer_name(printer_name)
    if not base:
        return []
    candidates = [base]
    try:
        import win32print  # type: ignore
        flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        for item in win32print.EnumPrinters(flags):
            try:
                pname = str(item[2]).strip()
            except Exception:
                continue
            if not pname:
                continue
            low_base = base.lower()
            low_p = pname.lower()
            if low_base != low_p and low_base not in low_p and low_p not in low_base:
                continue

            candidates.append(pname)
            try:
                h = win32print.OpenPrinter(pname)
                info = win32print.GetPrinter(h, 2)
                win32print.ClosePrinter(h)
                port = str(info.get("pPortName", "")).strip()
            except Exception:
                port = ""
            if port:
                candidates.extend([
                    f"{pname} on {port}:",
                    f"{pname} on {port}",
                    f"{pname} en {port}:",
                    f"{pname} en {port}",
                ])
    except Exception:
        pass
    return list(dict.fromkeys(candidates))


def _set_excel_active_printer(excel_app, printer_name: str) -> bool:
    candidates = _excel_active_printer_candidates(printer_name)
    # Excel suele mantener un sufijo de puerto en ActivePrinter; lo reutilizamos.
    try:
        current_ap = str(excel_app.ActivePrinter or "").strip()
        low = current_ap.lower()
        suffixes = []
        for sep in (" on ", " en "):
            idx = low.rfind(sep)
            if idx != -1 and current_ap.endswith(":"):
                suffixes.append(current_ap[idx:])
        if suffixes:
            expanded = []
            for c in candidates:
                expanded.append(c)
                for sfx in suffixes:
                    expanded.append(f"{c}{sfx}")
            candidates = expanded
    except Exception:
        pass

    for candidate in list(dict.fromkeys(candidates)):
        try:
            excel_app.ActivePrinter = candidate
            log_evento(f"[print] ActivePrinter aplicado: {candidate}", "info")
            return True
        except Exception:
            continue
    return False


def _imprimir_windows_printto(xlsx_path: Path, printer_name: str) -> bool:
    """
    Intenta enrutar explícitamente a una impresora en Windows sin LibreOffice.
    """
    target = _resolve_windows_printer_name(printer_name)
    if not target:
        return False
    try:
        import win32api  # type: ignore
        win32api.ShellExecute(0, "printto", str(xlsx_path), f'"{target}"', ".", 0)
        log_evento(f"Impresión por printto: {xlsx_path.name} -> {target}", "info")
        return True
    except Exception as e:
        log_evento(f"[print] printto falló para '{target}': {e}", "warning")
        return False


def generar_excel_temporal(df: pd.DataFrame, titulo: str, sheet_name: str = "Listado") -> Path:
    """
    Genera un .xlsx temporal con:
      - Título (fila 1) fusionado, negrita, centrado
      - Encabezados (fila 2)
      - Datos (desde fila 3)
      - Bordes finos y centrado
      - Autoajuste real de columnas (sin depender de util externa)
      - Impresión EN HORIZONTAL y ajuste a 1 página de ancho
    """
    if df is None or df.empty:
        raise ValueError("El DataFrame está vacío; no se puede generar Excel temporal.")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ncols = max(1, len(df.columns))

    # --- Título (fila 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    celda_titulo = ws.cell(row=1, column=1, value=titulo)
    celda_titulo.font = Font(bold=True, size=14)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

    # --- Encabezados (fila 2)
    for idx, col in enumerate(df.columns, start=1):
        c = ws.cell(row=2, column=idx, value=str(col))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Datos (desde fila 3)
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=value)
            c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Bordes finos (encabezados + datos)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for fila in ws.iter_rows(min_row=2, max_row=2 + len(df), min_col=1, max_col=ncols):
        for c in fila:
            c.border = thin_border

    # --- Autoajuste de columnas (robusto)
    # Heurística por longitud de cadena; padding y límites razonables.
    PAD = 2          # “aire” lateral
    MIN_W = 10       # ancho mínimo (excel units aprox)
    MAX_W = 100      # ancho máximo por columna
    for col_idx in range(1, ncols + 1):
        header = ws.cell(row=2, column=col_idx).value
        max_len = len(str(header)) if header is not None else 0

        # Revisa todas las filas de datos
        for r in range(3, 3 + len(df)):
            v = ws.cell(row=r, column=col_idx).value
            l = len(str(v)) if v is not None else 0
            if l > max_len:
                max_len = l

        # Calcula ancho aproximado (excel units ~ caracteres)
        width = max(MIN_W, min(MAX_W, max_len + PAD))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # --- Config de página: horizontal y a 1 página de ancho
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

    # --- Guardar temporal
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = Path(tmp.name)
    wb.save(str(temp_path))
    log_evento(f"Archivo temporal Excel generado: {temp_path}", "info")
    return temp_path


def enviar_a_impresora(archivo: Path, impresora_linux: Optional[str] = "Default", cleanup: bool = False) -> None:
    """
    Envía el .xlsx a la impresora por defecto del sistema.
    - Windows:   Excel COM; si falla usa mecanismos nativos de Windows.
    - Linux:     LibreOffice headless (--pt <impresora>), fallback a 'lp'.
    - macOS:     'lp'.

    Vars entorno:
      - EXCELCIOR_PRINTER      -> nombre impresora destino (Windows/Linux)
      - EXCELCIOR_PRINT_TIMEOUT-> segundos de timeout (int, default 25)
    """
    if not archivo or not Path(archivo).exists():
        raise FileNotFoundError(f"No existe el archivo a imprimir: {archivo}")

    sistema = platform.system()
    try:
        if sistema == "Windows":
            _imprimir_windows(archivo)
        elif sistema == "Linux":
            _imprimir_linux(archivo, impresora_linux=impresora_linux)
        elif sistema == "Darwin":
            _imprimir_macos(archivo)
        else:
            raise OSError(f"Sistema no soportado para impresión directa: {sistema}")
    except Exception as e:
        log_evento(f"Error al enviar a impresora: {e}", "error")
        raise
    finally:
        if cleanup:
            try:
                Path(archivo).unlink(missing_ok=True)
                log_evento(f"Temporal eliminado: {archivo}", "info")
            except Exception as ex:
                log_evento(f"No se pudo eliminar temporal: {archivo} ({ex})", "warning")


def enviar_a_impresora_configurable(
    archivo: Path,
    config: Optional[dict] = None,
    default_timeout_s: int = 120,
) -> None:
    """
    Adaptador único para impresión desde módulos de negocio.
    Lee impresora/timeout desde config y aplica compatibilidad de firmas.
    """
    cfg = config if isinstance(config, dict) else {}
    printer_name = cfg.get("printer_name") or cfg.get("printer") or cfg.get("impresora")

    timeout_s = default_timeout_s
    if "print_timeout_s" in cfg:
        try:
            timeout_s = int(cfg.get("print_timeout_s"))
        except Exception:
            timeout_s = default_timeout_s
    os.environ["EXCELCIOR_PRINT_TIMEOUT"] = str(timeout_s)

    if printer_name:
        os.environ["EXCELCIOR_PRINTER"] = str(printer_name)

    log_evento(
        f"[print] archivo='{Path(archivo).name}', printer='{printer_name or 'default-SO'}', timeout_s={timeout_s}",
        "info",
    )

    if printer_name:
        # 1) Firma estilo moderno (si existiera)
        try:
            return enviar_a_impresora(archivo, printer_name=printer_name)
        except TypeError:
            pass
        # 2) Firma posicional del proyecto actual: (archivo, impresora_linux, cleanup)
        try:
            return enviar_a_impresora(archivo, printer_name)
        except TypeError:
            pass
        # 3) Alias explícito de la firma actual
        try:
            return enviar_a_impresora(archivo, impresora_linux=printer_name)
        except TypeError:
            pass

    return enviar_a_impresora(archivo)


# ----------------- Helpers por SO -----------------

def _imprimir_windows(xlsx_path: Path) -> None:
    """
    Windows: intenta en orden:
      1) Excel COM (si Excel está instalado y registrado)
      2) ShellExecute 'print' como último recurso (sin LibreOffice)
    """
    # 1) Excel COM
    try:
        import pythoncom
        import win32print  # type: ignore
        from win32com.client import Dispatch  # pywin32

        pythoncom.CoInitialize()
        excel = None
        wb = None
        old_default_printer = None
        try:
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
            sh = wb.Worksheets(1)

            # Página horizontal y 1 página ancho
            sh.PageSetup.Orientation = 2         # xlLandscape
            sh.PageSetup.Zoom = False
            sh.PageSetup.FitToPagesWide = 1
            sh.PageSetup.FitToPagesTall = False

            used = sh.UsedRange
            used.Borders.LineStyle = 1
            used.HorizontalAlignment = -4108
            used.VerticalAlignment = -4108
            used.Columns.AutoFit()

            forced_printer = os.environ.get("EXCELCIOR_PRINTER", "").strip()
            if forced_printer:
                resolved = _resolve_windows_printer_name(forced_printer)
                try:
                    old_default_printer = win32print.GetDefaultPrinter()
                except Exception:
                    old_default_printer = None
                if old_default_printer and old_default_printer.strip().lower() != resolved.strip().lower():
                    try:
                        win32print.SetDefaultPrinter(resolved)
                        log_evento(f"[print] Default temporal (COM): {resolved}", "info")
                    except Exception as e:
                        log_evento(f"[print] No se pudo forzar default temporal: {e}", "warning")
                if not _set_excel_active_printer(excel, resolved):
                    raise RuntimeError(
                        f"No se pudo aplicar ActivePrinter explicito para '{resolved}' en Excel COM."
                    )
                try:
                    log_evento(f"[print] ActivePrinter final COM: {excel.ActivePrinter}", "info")
                except Exception:
                    pass

            sh.PrintOut()
            wb.Close(SaveChanges=False)
            log_evento(f"Impresión enviada por Excel COM: {xlsx_path.name}", "info")
            return
        finally:
            try:
                if old_default_printer:
                    win32print.SetDefaultPrinter(old_default_printer)
                    log_evento(f"[print] Default restaurada: {old_default_printer}", "info")
            except Exception:
                pass
            try:
                if wb:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    except Exception as com_err:
        log_evento(f"Excel COM no disponible o falló: {com_err}", "warning")

    # Si hay impresora forzada y COM no pudo respetarla, no intentar rutas ambiguas.
    forced_printer = os.environ.get("EXCELCIOR_PRINTER", "").strip()
    if forced_printer:
        if _imprimir_windows_printto(xlsx_path, forced_printer):
            return
        raise RuntimeError(
            f"No se pudo imprimir por Excel COM respetando la impresora forzada '{forced_printer}'."
        )

    # Último recurso: asociación de Windows
    try:
        os.startfile(str(xlsx_path), "print")
        log_evento(f"Impresión enviada por asociación de Windows: {xlsx_path.name}", "info")
        return
    except Exception as e:
        raise RuntimeError(
            "No se pudo imprimir: Excel COM no disponible y falló la asociación de Windows."
        ) from e


def _imprimir_linux(xlsx_path: Path, impresora_linux: Optional[str] = "Default") -> None:
    from subprocess import run, PIPE

    lo_cmd = [
        "libreoffice",
        "--headless",
        "--pt", impresora_linux or "Default",
        str(Path(xlsx_path).resolve()),
    ]
    log_evento(f"[Linux] Intentando LibreOffice: {' '.join(lo_cmd)}", "info")
    res = run(lo_cmd, stdout=PIPE, stderr=PIPE, text=True)
    if res.returncode != 0:
        log_evento(f"[Linux] LibreOffice falló ({res.returncode}). stderr: {res.stderr.strip()}", "warning")
        # Fallback a 'lp'
        lp_cmd = ["lp", str(Path(xlsx_path).resolve())]
        log_evento(f"[Linux] Intentando lp: {' '.join(lp_cmd)}", "info")
        res2 = run(lp_cmd, stdout=PIPE, stderr=PIPE, text=True)
        if res2.returncode != 0:
            log_evento(f"[Linux] 'lp' falló ({res2.returncode}). stderr: {res2.stderr.strip()}", "error")
            raise RuntimeError("No se pudo imprimir con LibreOffice ni con lp en Linux.")
        else:
            log_evento(f"[Linux] Enviado a impresora (lp): {xlsx_path.name}", "info")
    else:
        log_evento(f"[Linux] Enviado a impresora (LibreOffice): {xlsx_path.name}", "info")


def _imprimir_macos(xlsx_path: Path) -> None:
    from subprocess import run, PIPE
    lp_cmd = ["lp", str(Path(xlsx_path).resolve())]
    log_evento(f"[macOS] Imprimiendo con lp: {' '.join(lp_cmd)}", "info")
    res = run(lp_cmd, stdout=PIPE, stderr=PIPE, text=True)
    if res.returncode != 0:
        log_evento(f"[macOS] 'lp' falló ({res.returncode}). stderr: {res.stderr.strip()}", "error")
        raise RuntimeError("No se pudo imprimir con lp en macOS.")
    else:
        log_evento(f"[macOS] Enviado a impresora (lp): {xlsx_path.name}", "info")


# ---------- Descubrimiento de soffice/LibreOffice en Windows ----------

def _find_soffice_on_windows() -> Optional[str]:
    """Devuelve ruta a soffice.exe/.COM si se encuentra; None en caso contrario."""
    from shutil import which

    # Preferir .exe si hay en PATH
    path_exe = which("soffice")
    if path_exe:
        return path_exe

    # Rutas típicas
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files\LibreOffice\program\soffice.COM",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.COM",
    ]
    for c in candidates:
        if Path(c).exists():
            return c

    # Registro
    try:
        import winreg
        reg_paths = [
            (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\LibreOffice\UNO\InstallPath"),
            (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\LibreOffice\UNO\InstallPath"),
            (winreg.HKEY_CURRENT_USER, r"SOFTWARE\LibreOffice\UNO\InstallPath"),
        ]
        for hive, subkey in reg_paths:
            try:
                with winreg.OpenKey(hive, subkey) as k:
                    val, _ = winreg.QueryValueEx(k, "")
                    exe = Path(val) / "program" / "soffice.exe"
                    com = Path(val) / "program" / "soffice.COM"
                    if exe.exists():
                        return str(exe)
                    if com.exists():
                        return str(com)
            except FileNotFoundError:
                continue
    except Exception:
        pass

    return None


# ---------- Ejecución de soffice (o similar) con timeout ----------

def _imprimir_via_soffice_like(app_path: Path, xlsx_path: Path) -> None:
    """
    Lanza LibreOffice/soffice con timeout controlado para evitar cuelgues.
    Flags “silent” para no mostrar diálogos.
    Respeta EXCELCIOR_PRINTER y EXCELCIOR_PRINT_TIMEOUT.
    """
    import subprocess as sp

    # Normaliza a .exe si existe junto a .COM
    app_path = Path(app_path)
    if app_path.name.lower().endswith(".com"):
        exe_candidate = app_path.with_suffix(".exe")
        if exe_candidate.exists():
            app_path = exe_candidate

    printer = os.environ.get("EXCELCIOR_PRINTER", "Default")
    timeout_s = int(os.environ.get("EXCELCIOR_PRINT_TIMEOUT", "25"))

    cmd = [
        str(app_path),
        "--headless",
        "--invisible",
        "--norestore",
        "--nolockcheck",
        "--nodefault",
        "--nologo",
        "--nofirststartwizard",
        "--pt", printer,
        str(Path(xlsx_path).resolve()),
    ]

    # En Windows, evita abrir ventana de consola
    creationflags = 0
    startupinfo = None
    if platform.system() == "Windows":
        creationflags = 0x08000000  # CREATE_NO_WINDOW
        startupinfo = sp.STARTUPINFO()
        startupinfo.dwFlags |= sp.STARTF_USESHOWWINDOW

    log_evento(f"Intentando imprimir con: {' '.join(cmd)}", "info")

    try:
        proc = sp.Popen(
            cmd,
            stdout=sp.PIPE,
            stderr=sp.PIPE,
            text=True,
            creationflags=creationflags,
            startupinfo=startupinfo,
        )
        try:
            stdout, stderr = proc.communicate(timeout=timeout_s)
        except sp.TimeoutExpired:
            proc.kill()
            stdout, stderr = proc.communicate()
            msg = f"Tiempo de espera excedido ({timeout_s}s) imprimiendo con {app_path.name}."
            log_evento(msg + f" stderr: {str(stderr).strip()[:400]}", "error")
            raise RuntimeError(msg)

        rc = proc.returncode
        if rc != 0:
            # 3221225786 (0xC000013A) suele indicar interrupción/aborto
            log_evento(
                f"Impresión por '{app_path.name}' falló ({rc}). stderr: {str(stderr).strip()[:400]}",
                "error",
            )
            raise RuntimeError(f"No se pudo imprimir con {app_path.name}.")
        else:
            if stdout:
                log_evento(f"{app_path.name} stdout: {stdout.strip()[:200]}", "debug")
            log_evento(f"Enviado a impresora ({app_path.name}): {xlsx_path.name}", "info")

    except FileNotFoundError as e:
        raise RuntimeError(f"No se encontró ejecutable para impresión: {app_path}") from e


def convert_xlsx_to_pdf(xlsx_path: Path, output_dir: Optional[Path] = None) -> Path:
    """
    Convierte un .xlsx a .pdf con LibreOffice (soffice) y devuelve la ruta del PDF.
    """
    import subprocess as sp
    from shutil import which

    src = Path(xlsx_path)
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente para convertir: {src}")

    outdir = Path(output_dir) if output_dir else src.parent
    outdir.mkdir(parents=True, exist_ok=True)

    soffice = os.environ.get("EXCELCIOR_PRINT_APP", "").strip().strip('"')
    if soffice:
        soffice_path = Path(soffice)
    else:
        if platform.system() == "Windows":
            found = _find_soffice_on_windows()
            soffice_path = Path(found) if found else None
        else:
            found = which("soffice") or which("libreoffice")
            soffice_path = Path(found) if found else None

    if not soffice_path:
        raise RuntimeError(
            "No se encontró LibreOffice/soffice para convertir a PDF. "
            "Define EXCELCIOR_PRINT_APP o instala LibreOffice."
        )

    timeout_s = int(os.environ.get("EXCELCIOR_PRINT_TIMEOUT", "60"))
    cmd = [
        str(soffice_path),
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(outdir),
        str(src.resolve()),
    ]
    log_evento(f"Convirtiendo XLSX a PDF con: {' '.join(cmd)}", "info")

    creationflags = 0
    startupinfo = None
    if platform.system() == "Windows":
        creationflags = 0x08000000  # CREATE_NO_WINDOW
        startupinfo = sp.STARTUPINFO()
        startupinfo.dwFlags |= sp.STARTF_USESHOWWINDOW

    try:
        proc = sp.Popen(
            cmd,
            stdout=sp.PIPE,
            stderr=sp.PIPE,
            text=True,
            creationflags=creationflags,
            startupinfo=startupinfo,
        )
        stdout, stderr = proc.communicate(timeout=timeout_s)
    except sp.TimeoutExpired:
        proc.kill()
        raise RuntimeError(f"Tiempo de espera excedido ({timeout_s}s) al convertir a PDF.")

    if proc.returncode != 0:
        raise RuntimeError(
            f"Conversión a PDF falló ({proc.returncode}). "
            f"stderr: {str(stderr).strip()[:300]}"
        )

    pdf_path = outdir / f"{src.stem}.pdf"
    if not pdf_path.exists():
        raise RuntimeError(
            f"LibreOffice no generó el PDF esperado: {pdf_path}. "
            f"stdout: {str(stdout).strip()[:200]}"
        )

    log_evento(f"PDF generado correctamente: {pdf_path}", "info")
    return pdf_path


def enviar_pdf_a_impresora(pdf_path: Path, cleanup: bool = False) -> None:
    """
    Envía un PDF a impresión usando mecanismos nativos por plataforma.
    """
    from subprocess import run, PIPE

    pdf = Path(pdf_path)
    if not pdf.exists():
        raise FileNotFoundError(f"No existe el PDF a imprimir: {pdf}")

    try:
        if platform.system() == "Windows":
            os.startfile(str(pdf), "print")
            log_evento(f"PDF enviado a impresora (Windows shell): {pdf.name}", "info")
        else:
            cmd = ["lp", str(pdf.resolve())]
            res = run(cmd, stdout=PIPE, stderr=PIPE, text=True)
            if res.returncode != 0:
                raise RuntimeError(f"'lp' devolvió {res.returncode}: {res.stderr.strip()}")
            log_evento(f"PDF enviado a impresora (lp): {pdf.name}", "info")
    except Exception as e:
        log_evento(f"Error imprimiendo PDF: {e}", "error")
        raise
    finally:
        if cleanup:
            try:
                pdf.unlink(missing_ok=True)
            except Exception:
                pass
