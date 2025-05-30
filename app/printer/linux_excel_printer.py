from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import subprocess

def print_excel_linux(filepath: Path, mode: str):
    if not filepath.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    wb = load_workbook(filepath)
    sheet = wb.active

    fecha = datetime.now().strftime("%d/%m/%Y")
    titulo = {
        "fedex": f"FIN DE DÍA FEDEX - {fecha}",
        "urbano": f"FIN DE DÍA URBANO - {fecha}"
    }.get(mode, f"LISTADO GENERAL - {fecha}")

    # Insertar fila al inicio y colocar título
    sheet.insert_rows(1)
    sheet.cell(row=1, column=1).value = titulo
    sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

    # Bordes y alineación centrada
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    wb.save(filepath)

    # Opcional: imprimir automáticamente con LibreOffice
    try:
        subprocess.run(["libreoffice", "--headless", "--pt", "Nombre_de_Tu_Impresora", str(filepath)], check=True)
    except Exception as e:
        print(f"No se pudo imprimir automáticamente: {e}")
