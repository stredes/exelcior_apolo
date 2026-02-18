# app/printer/printer_etiquetas.py
# -*- coding: utf-8 -*-
"""
GeneraciÃ³n e impresiÃ³n de etiquetas desde plantilla Excel y soporte de impresiÃ³n PDF.
- Windows:
    â€¢ Primero intenta Excel COM (si hay Office).
    â€¢ Luego LibreOffice (soffice) directo a impresora (recomendado para evitar asociaciÃ³n).
    â€¢ Ãšltimo recurso: asociaciÃ³n del sistema (os.startfile(..., 'print')).
- Linux/macOS:
    â€¢ LibreOffice (soffice) o fallback 'lp'.
"""

from __future__ import annotations

import os
import platform
import subprocess as sp
import time
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Optional, Dict
from contextlib import contextmanager

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
import pandas as pd

from app.core.logger_eventos import log_evento

# ----------------- Imports condicionales (Windows) -----------------
try:
    if platform.system() == "Windows":
        import pythoncom  # pywin32
        from win32com.client import Dispatch  # type: ignore
    else:
        pythoncom = None  # type: ignore
        Dispatch = None   # type: ignore
except Exception:
    pythoncom = None  # type: ignore
    Dispatch = None   # type: ignore

# ----------------- Constantes / Config -----------------
TEMP_DIR = Path("temp")
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# Mapa de celdas por campo
CELDAS_MAP: Dict[str, str] = {
    "rut": "B2",
    "razsoc": "B3",
    "dir": "B4",
    "comuna": "B5",
    "guia": "B6",
    "bultos": "B7",
    "transporte": "B8",
}

# Impresora por defecto (puedes sobreescribir con EXCELCIOR_PRINTER)
DEFAULT_PRINTER = os.environ.get("EXCELCIOR_PRINTER", "").strip()

# Timeout en segundos para procesos de impresiÃ³n (LibreOffice)
PRINT_TIMEOUT_S = int(os.environ.get("EXCELCIOR_PRINT_TIMEOUT", "25"))

# Ejecutable forzado opcional (ruta a soffice)
FORCED_PRINT_APP = os.environ.get("EXCELCIOR_PRINT_APP", "").strip().strip('"')


# ----------------- Utilidades -----------------
def _ensure_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

def _find_soffice() -> Optional[str]:
    """
    Devuelve ruta a 'soffice' si estÃ¡ disponible. Busca en:
    - PATH
    - Rutas tÃ­picas de Windows
    - Registro de LibreOffice en Windows
    """
    from shutil import which

    exe = which("soffice") or which("libreoffice")
    if exe:
        return exe

    # Windows: rutas tÃ­picas
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files\LibreOffice\program\soffice.COM",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.COM",
    ]
    for c in candidates:
        if Path(c).exists():
            return c

    # Windows: registro
    if platform.system() == "Windows":
        try:
            import winreg  # type: ignore

            reg_paths = [
                (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\LibreOffice\UNO\InstallPath"),
                (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\LibreOffice\UNO\InstallPath"),
                (winreg.HKEY_CURRENT_USER, r"SOFTWARE\LibreOffice\UNO\InstallPath"),
            ]
            for hive, subkey in reg_paths:
                try:
                    with winreg.OpenKey(hive, subkey) as k:
                        val, _ = winreg.QueryValueEx(k, "")
                        exe = Path(val) / "program" / "soffice.exe"
                        if exe.exists():
                            return str(exe)
                        com = Path(val) / "program" / "soffice.COM"
                        if com.exists():
                            return str(com)
                except FileNotFoundError:
                    continue
        except Exception:
            pass

    return None

def _normalize_soffice(app: str) -> str:
    """Normaliza soffice.COM â†’ soffice.exe si existe el .exe al lado (Windows)."""
    p = Path(app)
    if p.suffix.lower() == ".com":
        exe = p.with_suffix(".exe")
        if exe.exists():
            return str(exe)
    return app

def _run_cmd(cmd: list[str], timeout_s: int = PRINT_TIMEOUT_S) -> None:
    """Ejecuta comando con timeout, loguea y lanza si rc != 0."""
    creationflags = 0
    startupinfo = None
    if platform.system() == "Windows":
        creationflags = 0x08000000  # CREATE_NO_WINDOW
        startupinfo = sp.STARTUPINFO()
        startupinfo.dwFlags |= sp.STARTF_USESHOWWINDOW

    log_evento(f"â–¶ Ejecutando: {' '.join(cmd)}", "info")
    try:
        proc = sp.Popen(
            cmd,
            stdout=sp.PIPE,
            stderr=sp.PIPE,
            text=True,
            creationflags=creationflags,
            startupinfo=startupinfo,
        )
        try:
            stdout, stderr = proc.communicate(timeout=timeout_s)
        except sp.TimeoutExpired:
            proc.kill()
            stdout, stderr = proc.communicate()
            log_evento(f"â³ Timeout ({timeout_s}s). stderr: {str(stderr).strip()[:400]}", "error")
            raise RuntimeError(f"Tiempo de espera excedido ({timeout_s}s).")

        if stdout:
            log_evento(stdout.strip()[:400], "debug")
        if proc.returncode != 0:
            log_evento(f"Comando fallÃ³ (rc={proc.returncode}). stderr: {str(stderr).strip()[:400]}", "error")
            raise RuntimeError(f"Error al ejecutar comando (rc={proc.returncode}).")
    except FileNotFoundError as e:
        raise RuntimeError(f"No se encontrÃ³ ejecutable: {cmd[0]}") from e


def _windows_printer_names() -> list[str]:
    if platform.system() != "Windows":
        return []
    try:
        import win32print  # type: ignore

        flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        salida = []
        for item in win32print.EnumPrinters(flags):
            try:
                nombre = str(item[2]).strip()
            except Exception:
                continue
            if nombre:
                salida.append(nombre)
        # Deduplicar preservando orden
        vistos = set()
        out = []
        for n in salida:
            if n not in vistos:
                out.append(n)
                vistos.add(n)
        return out
    except Exception:
        return []


def _resolve_windows_printer_name(alias: str) -> str:
    """
    Resuelve un alias (ej: 'URBANO') al nombre real de cola en Windows.
    """
    base = (alias or "").strip()
    if not base or platform.system() != "Windows":
        return base

    nombres = _windows_printer_names()
    if not nombres:
        return base

    lower = base.lower()
    for n in nombres:
        if n.lower() == lower:
            return n
    for n in nombres:
        if lower in n.lower():
            return n
    for n in nombres:
        if n.lower() in lower:
            return n
    return base


@contextmanager
def _temporary_default_printer(printer_name: str):
    """
    Establece temporalmente la impresora predeterminada en Windows y luego la restaura.
    """
    if platform.system() != "Windows" or not printer_name:
        yield
        return
    try:
        import win32print  # type: ignore

        old_default = win32print.GetDefaultPrinter()
        target = _resolve_windows_printer_name(printer_name)
        if target:
            win32print.SetDefaultPrinter(target)
            log_evento(f"ðŸ–¨ï¸ Default temporal: {target}", "info")
        try:
            yield
        finally:
            if old_default:
                win32print.SetDefaultPrinter(old_default)
                log_evento(f"â†©ï¸ Default restaurada: {old_default}", "info")
    except Exception:
        # Si no se puede cambiar default, continuar sin bloquear.
        yield


def _imprimir_windows_asociacion(file_path: Path, printer: str) -> None:
    """
    Fallback Windows sin Excel/LibreOffice:
    1) Intenta 'printto' directo a impresora.
    2) Si falla, usa 'print' con default temporal y espera breve antes de restaurar.
    """
    if platform.system() != "Windows":
        raise RuntimeError("Asociacion Windows solo aplica en Windows.")

    target = _resolve_windows_printer_name(printer or "")
    if target:
        try:
            import win32api  # type: ignore

            win32api.ShellExecute(0, "printto", str(file_path), f'"{target}"', ".", 0)
            log_evento(f"ImpresiÃ³n por asociaciÃ³n Windows (printto): {file_path.name} -> {target}", "info")
            return
        except Exception as e:
            log_evento(f"printto fallÃ³ para '{target}': {e}", "warning")

    with _temporary_default_printer(target):
        os.startfile(str(file_path), "print")  # type: ignore
        # startfile retorna antes de que se enrute realmente el trabajo.
        time.sleep(4)
    log_evento(f"ImpresiÃ³n por asociaciÃ³n Windows: {file_path.name}", "info")


def _excel_printer_candidates(nombre_impresora: str) -> list[str]:
    """
    Construye variantes para Excel.ActivePrinter.
    Excel en Windows suele exigir formato: '<Nombre> on <Puerto>:'.
    """
    base = (nombre_impresora or "").strip()
    if not base or platform.system() != "Windows":
        return [base] if base else []

    candidatos = [base]
    resolved = _resolve_windows_printer_name(base)
    if resolved:
        candidatos.insert(0, resolved)
    try:
        import win32print  # type: ignore

        flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        for item in win32print.EnumPrinters(flags):
            try:
                pname = str(item[2]).strip()
            except Exception:
                continue
            if not pname:
                continue
            if resolved and resolved.lower() not in pname.lower() and pname.lower() not in resolved.lower() and base.lower() not in pname.lower() and pname.lower() not in base.lower():
                continue

            candidatos.append(pname)
            try:
                h = win32print.OpenPrinter(pname)
                info = win32print.GetPrinter(h, 2)
                win32print.ClosePrinter(h)
                port = str(info.get("pPortName", "")).strip()
            except Exception:
                port = ""

            if port:
                # Excel es sensible al formato exacto; probar variantes.
                candidatos.append(f"{pname} on {port}:")
                candidatos.append(f"{pname} on {port}")
                candidatos.append(f"{pname} en {port}:")
                candidatos.append(f"{pname} en {port}")
    except Exception:
        pass

    # Deduplicar preservando orden
    vistos = set()
    salida = []
    for c in candidatos:
        if c and c not in vistos:
            salida.append(c)
            vistos.add(c)
    return salida


# ----------------- GeneraciÃ³n de etiqueta (xlsx) -----------------
def generar_etiqueta_excel(data: dict, output_path: Path) -> Path:
    """
    Genera la etiqueta XLSX directamente desde cÃ³digo (sin plantilla externa).
    """
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Etiqueta"

        # Estilo base (bordes visibles en impresión)
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        label_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
        # Ajuste visual: fuente y alturas mayores para ocupar mejor la etiqueta 10x14.
        label_font = Font(name="Calibri", size=15, bold=True, color="111827")
        value_font = Font(name="Calibri", size=16, bold=True, color="111827")
        center = Alignment(vertical="center")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 38

        field_labels = {
            "rut": "RUT",
            "razsoc": "Cliente",
            "dir": "Direccion",
            "comuna": "Comuna",
            "guia": "Guia",
            "bultos": "Bultos",
            "transporte": "Transporte",
        }

        for campo, celda in CELDAS_MAP.items():
            row = ws[celda].row
            label_cell = ws[f"A{row}"]
            value_cell = ws[celda]
            label_cell.value = field_labels.get(campo, campo.title())
            value_cell.value = data.get(campo, "")

            label_cell.fill = label_fill
            label_cell.font = label_font
            value_cell.font = value_font
            label_cell.border = border
            value_cell.border = border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.alignment = center
            ws.row_dimensions[row].height = 42

        ws.merge_cells("A1:B1")
        header = ws["A1"]
        header.value = "Bodega Amilab\nEtiqueta de Despacho"
        header.font = Font(name="Calibri", size=18, bold=True, color="111827")
        # Header sin fondo, segÃºn requerimiento.
        header.fill = PatternFill(fill_type=None)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = border
        ws.row_dimensions[1].height = 54

        # Footer con fecha/hora de impresiÃ³n
        ws.merge_cells("A9:B9")
        footer = ws["A9"]
        footer.value = f"Impresion: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        footer.font = Font(name="Calibri", size=11, bold=False, color="374151")
        footer.alignment = Alignment(horizontal="right", vertical="center")
        footer.border = border
        ws.row_dimensions[9].height = 24

        # Bordes completos para toda el área imprimible, con contorno exterior reforzado.
        min_r, max_r, min_c, max_c = 1, 9, 1, 2
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = Border(
                    left=medium if c == min_c else thin,
                    right=medium if c == max_c else thin,
                    top=medium if r == min_r else thin,
                    bottom=medium if r == max_r else thin,
                )

        # Config de pagina 10x14 cm
        try:
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_margins = PageMargins(
                left=0.2, right=0.2, top=0.3, bottom=0.3, header=0.1, footer=0.1
            )
            ws.page_setup.paperWidth = "10cm"
            ws.page_setup.paperHeight = "14cm"
            if hasattr(ws, "sheet_properties") and hasattr(ws.sheet_properties, "pageSetUpPr"):
                ws.sheet_properties.pageSetUpPr.fitToPage = True  # type: ignore[attr-defined]
            ws.print_area = "A1:B9"
        except Exception as e:
            log_evento(f"âš ï¸ No se pudo aplicar tamano 10x14 cm: {e}", "warning")

        wb.save(output_path)
        log_evento(f"ðŸ“„ Etiqueta generada: {output_path}", "info")
        return output_path

    except Exception as e:
        log_evento(f"âŒ Error al generar etiqueta Excel: {e}", "error")
        raise RuntimeError(f"Error al generar etiqueta: {e}")


# ----------------- ImpresiÃ³n XLSX -----------------
def _imprimir_excel_windows_via_com(xlsx_path: Path, impresora: str | None) -> None:
    """Intenta imprimir con Excel COM en Windows."""
    if pythoncom is None or Dispatch is None:
        raise RuntimeError("COM/pywin32 no disponible en este sistema.")

    try:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
            hoja = wb.Sheets(1)

            # Respetar ajuste a pÃ¡gina
            hoja.PageSetup.Zoom = False
            hoja.PageSetup.FitToPagesWide = 1
            hoja.PageSetup.FitToPagesTall = 1

            if impresora:
                # Toma sufijo de puerto del ActivePrinter actual de Excel (el mÃ¡s confiable)
                suffixes = []
                try:
                    current_ap = str(excel.ActivePrinter or "").strip()
                    low = current_ap.lower()
                    for sep in (" on ", " en "):
                        idx = low.rfind(sep)
                        if idx != -1 and current_ap.endswith(":"):
                            suffixes.append(current_ap[idx:])
                except Exception:
                    pass

                aplicado = False
                ultimo_error = None
                candidates = _excel_printer_candidates(impresora)
                if suffixes:
                    expanded = []
                    for c in candidates:
                        expanded.append(c)
                        for sfx in suffixes:
                            expanded.append(f"{c}{sfx}")
                    candidates = expanded

                log_evento(
                    f"ðŸ§­ Candidatos ActivePrinter para '{impresora}': {candidates[:20]}",
                    "debug",
                )
                for candidato in candidates:
                    try:
                        excel.ActivePrinter = candidato
                        aplicado = True
                        log_evento(
                            f"ðŸ–¨ï¸ ActivePrinter aplicado: {candidato}",
                            "info",
                        )
                        break
                    except Exception as e:
                        ultimo_error = e
                        continue
                if not aplicado:
                    raise RuntimeError(
                        f"No se pudo seleccionar impresora '{impresora}' en Excel COM."
                    ) from ultimo_error
            hoja.PrintOut()

            log_evento(f"ðŸ–¨ï¸ Excel COM: {xlsx_path.name} -> {impresora or '[predeterminada]'}", "info")
        finally:
            try:
                if wb:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    except Exception as e:
        raise RuntimeError(f"Excel COM fallÃ³: {e}")

def _imprimir_via_soffice_xlsx(xlsx_path: Path, impresora: str | None) -> None:
    """
    Imprime XLSX con LibreOffice (soffice).
    """
    app = _normalize_soffice(FORCED_PRINT_APP or (_find_soffice() or ""))
    if not app:
        raise RuntimeError("No se encontrÃ³ LibreOffice (soffice). InstÃ¡lalo o define EXCELCIOR_PRINT_APP.")

    cmd = [
        app,
        "--headless", "--invisible", "--norestore", "--nolockcheck",
        "--nodefault", "--nologo", "--nofirststartwizard",
        "--pt", impresora or DEFAULT_PRINTER or "",
        str(xlsx_path.resolve()),
    ]
    _run_cmd(cmd, timeout_s=PRINT_TIMEOUT_S)
    log_evento(f"ðŸ–¨ï¸ soffice (xlsx): {xlsx_path.name} -> {impresora or DEFAULT_PRINTER or '[predeterminada]'}", "info")

def _imprimir_via_lp(file_path: Path) -> None:
    """Fallback bÃ¡sico en Linux/macOS usando 'lp'."""
    cmd = ["lp", str(file_path.resolve())]
    _run_cmd(cmd, timeout_s=PRINT_TIMEOUT_S)
    log_evento(f"ðŸ–¨ï¸ lp: {file_path.name}", "info")

def imprimir_excel(path: Path, impresora: Optional[str] = None) -> None:
    """
    Envia el .xlsx a imprimir:
      - Windows: Excel COM â†’ soffice â†’ asociaciÃ³n Windows (Ãºltimo recurso)
      - Linux/macOS: soffice â†’ lp
    """
    _ensure_exists(path)
    so = platform.system()
    printer = (impresora or DEFAULT_PRINTER).strip()
    if so == "Windows" and printer:
        printer = _resolve_windows_printer_name(printer)

    if so == "Windows":
        # 1) Excel COM
        try:
            _imprimir_excel_windows_via_com(path, printer or None)
            return
        except Exception as com_err:
            log_evento(f"Excel COM no disponible o fallÃ³: {com_err}", "warning")

        # 2) soffice
        try:
            _imprimir_via_soffice_xlsx(path, printer or None)
            return
        except Exception as lo_err:
            log_evento(f"LibreOffice no disponible o fallÃ³: {lo_err}", "warning")

        # 3) AsociaciÃ³n del sistema (puede fallar si no hay visor predeterminado)
        try:
            _imprimir_windows_asociacion(path, printer)
            return
        except Exception as e:
            raise RuntimeError(
                "No se pudo imprimir en Windows: COM fallÃ³ y no se encontrÃ³ LibreOffice. "
                "Instala Excel o LibreOffice, o define EXCELCIOR_PRINT_APP con la ruta a soffice.exe."
            ) from e

    else:
        # Linux / macOS
        try:
            _imprimir_via_soffice_xlsx(path, printer or None)
            return
        except Exception as lo_err:
            log_evento(f"LibreOffice no disponible o fallÃ³: {lo_err}", "warning")

        try:
            _imprimir_via_lp(path)
            return
        except Exception as e:
            raise RuntimeError(
                "No se pudo imprimir en este sistema: LibreOffice y 'lp' fallaron. "
                "Instala LibreOffice o configura CUPS correctamente."
            ) from e


# ----------------- ImpresiÃ³n PDF (nuevo) -----------------
def imprimir_pdf(path: Path, impresora: Optional[str] = None) -> None:
    """
    Imprime un PDF **sin** depender de la app predeterminada de Windows,
    usando LibreOffice cuando sea posible.
    """
    _ensure_exists(path)
    so = platform.system()
    printer = (impresora or DEFAULT_PRINTER).strip()

    # Intentar siempre con LibreOffice (evita WinError 1155)
    try:
        app = _normalize_soffice(FORCED_PRINT_APP or (_find_soffice() or ""))
        if app:
            cmd = [
                app, "--headless", "--invisible", "--norestore", "--nolockcheck",
                "--nodefault", "--nologo", "--nofirststartwizard",
                "--pt", printer or "",
                str(path.resolve()),
            ]
            _run_cmd(cmd, timeout_s=PRINT_TIMEOUT_S)
            log_evento(f"ðŸ–¨ï¸ soffice (pdf): {path.name} -> {printer or '[predeterminada]'}", "info")
            return
    except Exception as lo_err:
        log_evento(f"LibreOffice no disponible o fallÃ³ (pdf): {lo_err}", "warning")

    # Fallbacks por SO
    if so == "Windows":
        try:
            os.startfile(str(path), "print")  # type: ignore
            log_evento(f"ImpresiÃ³n por asociaciÃ³n Windows (pdf): {path.name}", "info")
            return
        except Exception as e:
            raise RuntimeError(
                "No se pudo imprimir PDF en Windows: LibreOffice y asociaciÃ³n fallaron. "
                "Instala LibreOffice o un visor PDF y configÃºralo como predeterminado."
            ) from e
    else:
        try:
            _imprimir_via_lp(path)
            return
        except Exception as e:
            raise RuntimeError(
                "No se pudo imprimir PDF en este sistema: LibreOffice y 'lp' fallaron."
            ) from e


def print_etiquetas(file_path, config, df: pd.DataFrame) -> None:
    """
    Imprime una etiqueta por cada fila del DataFrame (cada fila -> una etiqueta).
    Usa archivos temporales distintos para evitar locks.
    """
    try:
        if df is None or df.empty:
            raise ValueError("El DataFrame de etiquetas estÃ¡ vacÃ­o.")


        cfg = config if isinstance(config, dict) else {}
        configured_label_printer = (
            cfg.get("label_printer_name")
            or cfg.get("impresora_etiquetas")
            or cfg.get("label_printer")
            or cfg.get("printer_name")
            or ""
        )

        for _, row in df.iterrows():
            data = {
                "rut": row.get("RUT", ""),
                "razsoc": row.get("RazÃ³n Social", "") or row.get("Razon Social", ""),
                "dir": row.get("DirecciÃ³n", "") or row.get("Direccion", ""),
                "comuna": row.get("Comuna", ""),
                "guia": row.get("GuÃ­a", "") or row.get("Guia", ""),
                "bultos": row.get("Bultos", ""),
                "transporte": row.get("Transporte", "") or DEFAULT_PRINTER or "",
            }
            log_evento(f"ðŸ§¾ Generando etiqueta para: {data}", "info")

            with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=TEMP_DIR) as tmp:
                tmp_path = Path(tmp.name)

            generar_etiqueta_excel(data, tmp_path)
            imprimir_excel(tmp_path, configured_label_printer or data.get("transporte") or DEFAULT_PRINTER or None)

        log_evento("âœ… ImpresiÃ³n de todas las etiquetas finalizada.", "info")

    except Exception as e:
        log_evento(f"âŒ Error en impresiÃ³n mÃºltiple de etiquetas: {e}", "error")
        raise RuntimeError(f"Error en impresiÃ³n mÃºltiple de etiquetas: {e}")
