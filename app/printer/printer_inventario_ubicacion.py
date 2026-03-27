# Módulo: printer_inventario_ubicacion.py
# Descripción: Impresión automática del inventario filtrado por ubicación.

from __future__ import annotations

import os
import platform
import subprocess
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.core.logger_eventos import log_evento


def print_inventario_ubicacion(file_path=None, config=None, df: pd.DataFrame = None):
    """
    Entrada estandar para printer_map. Genera archivo Excel con estilo y lo imprime automaticamente.
    Compatible con llamada antigua: print_inventario_ubicacion(df=...).
    """
    temp_path: Path | None = None
    try:
        if isinstance(file_path, pd.DataFrame) and df is None:
            df = file_path
        if df is None:
            raise ValueError("No se recibio DataFrame para impresion de inventario por ubicacion.")
        if df.empty:
            raise ValueError("El DataFrame del inventario por ubicacion esta vacio.")

        fecha = datetime.now().strftime("%d/%m/%Y")
        titulo = f"INVENTARIO POR UBICACION - {fecha}"

        df_to_export = pd.DataFrame(columns=df.columns)
        df_to_export.loc[0] = [""] * len(df.columns)
        df_to_export = pd.concat([df_to_export, df], ignore_index=True)

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = Path(temp_file.name)

        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            df_to_export.to_excel(writer, index=False, sheet_name="Inventario")
            sheet = writer.book["Inventario"]

            total_columnas = len(df.columns)
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas)

            cell = sheet.cell(row=1, column=1)
            cell.value = titulo
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")

            borde_fino = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_idx in range(1, total_columnas + 1):
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].auto_size = True

            for row in sheet.iter_rows(min_row=2, max_row=df.shape[0] + 2, min_col=1, max_col=total_columnas):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = borde_fino

            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_margins.left = 0.2
            sheet.page_margins.right = 0.2
            sheet.page_margins.top = 0.3
            sheet.page_margins.bottom = 0.3
            sheet.page_margins.header = 0.1
            sheet.page_margins.footer = 0.1
            sheet.print_options.horizontalCentered = True

        log_evento(f"Archivo temporal generado para inventario por ubicacion: {temp_path}", "info")
        _enviar_a_impresora(temp_path)
        log_evento("Impresion de inventario por ubicacion completada correctamente.", "info")

    except Exception as e:
        log_evento(f"Error al imprimir inventario por ubicacion: {e}", "error")
        raise RuntimeError(f"Error al imprimir inventario por ubicacion: {e}")
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except Exception as cleanup_error:
                log_evento(f"No se pudo eliminar temporal de inventario por ubicacion: {cleanup_error}", "warning")


def _enviar_a_impresora(file_path: Path):
    sistema = platform.system()
    try:
        if sistema == "Windows":
            os.startfile(str(file_path), "print")
        elif sistema == "Linux":
            subprocess.run(["libreoffice", "--headless", "--pt", "Default", str(file_path)], check=True)
        elif sistema == "Darwin":
            subprocess.run(["lp", str(file_path)], check=True)
        else:
            raise OSError("Sistema operativo no compatible para impresion automatica.")
    except Exception as e:
        log_evento(f"Error al imprimir archivo: {e}", "error")
        raise RuntimeError(f"Error al enviar a impresora: {e}")
