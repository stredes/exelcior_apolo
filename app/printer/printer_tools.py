# app/printer/printer_tools.py
# -*- coding: utf-8 -*-
"""
Herramientas de impresión y preparación de datos para FedEx/Urbano.

• prepare_fedex_dataframe(df)   -> DataFrame normalizado y consolidado (agregación flexible de BULTOS)
• prepare_urbano_dataframe(df)  -> DataFrame normalizado y total PIEZAS
• insertar_bloque_firma_ws(ws)  -> Bloque "Nombre/Firma" al final de la hoja
• agregar_footer_info_ws(ws,t)  -> Pie con timestamp y total de piezas
• formatear_tabla_ws(ws)        -> Estilo de tabla profesional (anchos/bordes auto por modo)

Novedades (FedEx):
- Consolidación por masterTrackingNumber (no pieceTrackingNumber).
- Usa numberOfPackages del envío; NO suma por cada pieza (evita inflar/deflactar).
- EXCELCIOR_FEDEX_BULTOS_AGG = smart|max|min|last|sum (default: smart)
  • smart: si hay algún valor >= 2 en el grupo, usa el máximo; si no, toma el último.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Tuple, Optional

import numpy as np
import pandas as pd


# ======================================================================
#                      Normalización / Limpieza de datos
# ======================================================================

def _df_safe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    df2 = df2.replace({pd.NA: "", np.nan: ""})
    df2 = df2.replace({"nan": "", "<NA>": ""})
    for c in df2.columns:
        if pd.api.types.is_string_dtype(df2[c]) or pd.api.types.is_object_dtype(df2[c]):
            df2[c] = df2[c].astype(object)
    return df2


def _clean_text_series(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series([], dtype="string")
    s = series.astype("string").fillna("").str.strip()
    s = s.replace(to_replace=r"(?i)^(nan|<na>|none|null)$", value="", regex=True)
    return s


def _stringify_tracking(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series([], dtype="string")
    as_num = pd.to_numeric(series, errors="coerce")
    out = pd.Series("", index=series.index, dtype="string")
    mask_num = as_num.notna()
    if mask_num.any():
        out.loc[mask_num] = as_num.loc[mask_num].astype("Int64").astype("string")
    mask_rest = ~mask_num
    if mask_rest.any():
        out.loc[mask_rest] = series.loc[mask_rest].astype("string")
    out = out.fillna("").str.strip().replace({"nan": "", "<NA>": ""})
    return out


def _stringify_generic(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series([], dtype="string")
    as_num = pd.to_numeric(series, errors="coerce")
    out = pd.Series("", index=series.index, dtype="string")
    mask_num = as_num.notna()
    if mask_num.any():
        out.loc[mask_num] = as_num.loc[mask_num].astype("Int64").astype("string")
    mask_rest = ~mask_num
    if mask_rest.any():
        out.loc[mask_rest] = series.loc[mask_rest].astype("string")
    out = out.fillna("").str.strip().replace({"nan": "", "<NA>": ""})
    return out


def _normalize_date(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series([], dtype="string")
    dt = pd.to_datetime(series, errors="coerce", utc=False)
    mask_nat = dt.isna()
    if mask_nat.any():
        as_num = pd.to_numeric(series[mask_nat], errors="coerce")
        mask_num = as_num.notna()
        if mask_num.any():
            idx = as_num[mask_num].index
            dt2 = pd.to_datetime(as_num[mask_num], unit="D", origin="1899-12-30", errors="coerce")
            dt.loc[idx] = dt2
    out = pd.Series("", index=series.index, dtype="string")
    mask_ok = dt.notna()
    if mask_ok.any():
        out.loc[mask_ok] = dt.loc[mask_ok].dt.strftime("%Y-%m-%d")
    return out


# ======================================================================
#                  Utilidades de resolución de columnas
# ======================================================================

def _cimap(df: pd.DataFrame) -> dict:
    return {str(c).strip().lower(): c for c in df.columns}

def _pick_ci(df: pd.DataFrame, *names: str) -> Optional[str]:
    cmap = _cimap(df)
    for n in names:
        key = str(n).strip().lower()
        if key in cmap:
            return cmap[key]
    return None


# ======================================================================
#                  Agregación robusta para BULTOS (FedEx)
# ======================================================================

def _agg_bultos(series: pd.Series) -> int:
    """
    Modo por ENV (default: 'smart'):
      - smart: si hay algún valor >=2 en el grupo, devuelve el máximo; si no, el último.
      - max/min/last/sum: forzados.
    """
    mode = os.environ.get("EXCELCIOR_FEDEX_BULTOS_AGG", "smart").lower()
    b = pd.to_numeric(series, errors="coerce").fillna(0).astype(int)
    b.loc[b <= 0] = 1
    if b.empty:
        return 0

    if mode == "sum":
        return int(b.sum())
    if mode == "max":
        return int(b.max())
    if mode == "min":
        return int(b.min())
    if mode == "last":
        return int(b.iloc[-1])

    # smart (por defecto)
    if (b >= 2).any():
        return int(b.max())
    return int(b.iloc[-1])


# ------------------------- FEDEX -------------------------

def prepare_fedex_dataframe(df: pd.DataFrame) -> Tuple[pd.DataFrame, str, int]:
    """
    Devuelve (df_limpio, id_col, total_piezas).

    Reglas clave:
    - Consolida por masterTrackingNumber (ID del envío). Si no existe, cae a Tracking Number genérico.
    - Usa numberOfPackages (o alias) como BULTOS del envío; NO suma por cada pieza.
    - Column matching case-insensitive + strip.
    - Agregación _agg_bultos dentro del grupo (default 'smart').
    """
    cols_final = ["Tracking Number", "Fecha", "Referencia", "Ciudad", "Receptor", "BULTOS"]
    if df is None or df.empty:
        return pd.DataFrame(columns=cols_final), "", 0

    df = df.copy()
    cmap = _cimap(df)

    # ----------------- CASO 1: DF ya trae columnas finales -----------------
    have_all = all(str(h).strip().lower() in cmap for h in cols_final)
    if have_all:
        real_cols = [cmap[str(h).strip().lower()] for h in cols_final]
        out = df.loc[:, real_cols].copy()
        out.columns = cols_final

        # Prioriza columna oficial de paquetes si existe en el DF original
        b_alias_col = _pick_ci(
            df,
            "numberOfPackages", "numberofpackages",
            "totalPackageCount", "packageCount", "packagesCount", "totalPackages",
            "pieceCount", "packages", "pieces",
        )
        if b_alias_col is not None:
            b_alias = pd.to_numeric(df[b_alias_col], errors="coerce")
            out["BULTOS"] = b_alias.where(b_alias > 0, np.nan).fillna(
                pd.to_numeric(out["BULTOS"], errors="coerce")
            )

        # Normalización
        out["Tracking Number"] = _stringify_tracking(out["Tracking Number"])
        out["Fecha"]           = _normalize_date(out["Fecha"])
        out["Referencia"]      = _stringify_generic(out["Referencia"])
        out["Ciudad"]          = _clean_text_series(out["Ciudad"])
        out["Receptor"]        = _clean_text_series(out["Receptor"])

        b = pd.to_numeric(out["BULTOS"], errors="coerce").fillna(0).astype(int)
        b.loc[b <= 0] = 1
        out["BULTOS"] = b

        out = out.sort_values(["Tracking Number", "Fecha"], kind="stable")
        grouped = (
            out.groupby("Tracking Number", as_index=False)
               .agg({
                   "Fecha": "last",
                   "Referencia": "last",
                   "Ciudad": "last",
                   "Receptor": "last",
                   "BULTOS": _agg_bultos,
               })
        )
        total_piezas = int(grouped["BULTOS"].sum())
        grouped = _df_safe_for_excel(grouped)
        return grouped.reset_index(drop=True), "Tracking Number", total_piezas

    # --------------- CASO 2: Mapear alias y construir salida ---------------
    id_col_master = _pick_ci(df, "masterTrackingNumber", "mastertrackingnumber", "master tracking number")
    id_col_generic = _pick_ci(df, "trackingNumber", "tracking number", "Tracking Number", "tracking")
    id_col = id_col_master or id_col_generic
    if id_col is None:
        return pd.DataFrame(columns=cols_final), "", 0

    b_col = _pick_ci(
        df,
        "numberOfPackages", "numberofpackages",
        "totalPackageCount", "packageCount", "packagesCount", "totalPackages",
        "BULTOS", "bultos", "PIEZAS", "piezas",
        "pieces", "pieceCount", "packages"
    )
    if b_col:
        b = pd.to_numeric(df[b_col], errors="coerce").fillna(0).astype(int)
        b.loc[b <= 0] = 1
    else:
        b = pd.Series(1, index=df.index, dtype=int)

    fecha_col = _pick_ci(df, "shipDate", "Ship Date", "Fecha", "fecha", "date", "Date")
    ref_col   = _pick_ci(df, "reference", "Reference", "Referencia", "referencia", "Ref", "ref")
    city_col  = _pick_ci(df, "recipientCity", "recipient_city", "city", "City", "Ciudad", "ciudad", "Localidad", "localidad")
    recv_col  = _pick_ci(df, "recipientContactName", "recipientName", "recipient_name", "Receptor", "receptor", "Destinatario", "destinatario")

    base = pd.DataFrame(index=df.index)
    base["Tracking Number"] = _stringify_tracking(df[id_col])
    base["Fecha"]           = _normalize_date(df[fecha_col]) if fecha_col else pd.Series("", index=df.index, dtype="string")
    base["Referencia"]      = _stringify_generic(df[ref_col]) if ref_col else pd.Series("", index=df.index, dtype="string")
    base["Ciudad"]          = _clean_text_series(df[city_col]) if city_col else pd.Series("", index=df.index, dtype="string")
    base["Receptor"]        = _clean_text_series(df[recv_col]) if recv_col else pd.Series("", index=df.index, dtype="string")
    base["BULTOS"]          = b

    mask = base["Tracking Number"] != ""
    if recv_col:
        mask &= base["Receptor"] != ""
    base = base.loc[mask].copy()
    if base.empty:
        return pd.DataFrame(columns=cols_final), id_col, 0

    group_key = "Tracking Number"
    base = base.sort_values([group_key, "Fecha"], kind="stable")

    def _first_non_empty_last(s: pd.Series) -> str:
        s2 = s.fillna("").astype("string")
        nz = s2[s2.str.strip() != ""]
        return nz.iloc[-1] if not nz.empty else (s2.iloc[-1] if len(s2) else "")

    grouped = (
        base.groupby(group_key, as_index=False)
            .agg({
                "Fecha": "last",
                "Referencia": _first_non_empty_last,
                "Ciudad": _first_non_empty_last,
                "Receptor": _first_non_empty_last,
                "BULTOS": _agg_bultos,   # smart|max|min|last|sum
            })
    )

    _sd = pd.to_datetime(grouped["Fecha"], errors="coerce")
    grouped = grouped.assign(_sd=_sd).sort_values(["_sd", group_key], na_position="last").drop(columns=["_sd"])

    total_piezas = int(grouped["BULTOS"].sum())
    grouped = _df_safe_for_excel(grouped)
    return grouped.reset_index(drop=True), id_col, total_piezas


# ------------------------- URBANO -------------------------

def prepare_urbano_dataframe(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    cols_final = ["GUIA", "CLIENTE", "LOCALIDAD", "CIUDAD", "PIEZAS", "COD RASTREO"]
    if df is None or df.empty:
        return pd.DataFrame(columns=cols_final), 0

    df = df.copy()

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    guia_col  = pick("GUIA", "guia", "Guia")
    cli_col   = pick("CLIENTE", "cliente", "Cliente")
    loc_col   = pick("LOCALIDAD", "localidad", "Localidad")
    city_col  = pick("CIUDAD", "ciudad", "Ciudad")
    piezas_c  = pick("PIEZAS", "piezas", "Piezas", "BULTOS", "bultos")
    rastreo_c = pick("COD RASTREO", "COD_RASTREO", "codRastreo", "TRACKING", "tracking")

    out = pd.DataFrame(index=df.index)
    txt_guia  = _clean_text_series(df[guia_col])  if guia_col  else pd.Series("", index=df.index, dtype="string")
    txt_cli   = _clean_text_series(df[cli_col])   if cli_col   else pd.Series("", index=df.index, dtype="string")
    txt_loc   = _clean_text_series(df[loc_col])   if loc_col   else pd.Series("", index=df.index, dtype="string")
    txt_city  = _clean_text_series(df[city_col])  if city_col  else txt_loc.copy()
    # Evita imprimir rastreos como 61396.0; los normaliza a entero-texto.
    txt_track = _stringify_tracking(df[rastreo_c]) if rastreo_c else pd.Series("", index=df.index, dtype="string")

    out["GUIA"]        = txt_guia
    out["CLIENTE"]     = txt_cli
    out["LOCALIDAD"]   = txt_loc
    out["CIUDAD"]      = txt_city
    out["COD RASTREO"] = txt_track

    if piezas_c:
        raw_piezas = df[piezas_c]
        numeric = pd.to_numeric(raw_piezas, errors="coerce")

        if numeric.isna().any():
            extracted = (
                raw_piezas.astype(str)
                .str.replace(",", ".", regex=False)
                .str.extract(r"(\d+\.?\d*)")[0]
            )
            fallback_numeric = pd.to_numeric(extracted, errors="coerce")
            numeric = numeric.fillna(fallback_numeric)

        numeric = numeric.fillna(0)
        numeric = numeric.clip(lower=0)
        p = numeric.round().astype(int)
    else:
        p = pd.Series(1, index=df.index, dtype=int)
    out["PIEZAS"]      = p

    has_any_text = out[["GUIA", "CLIENTE", "LOCALIDAD", "CIUDAD", "COD RASTREO"]].apply(lambda s: s.str.strip() != "", axis=0).any(axis=1)
    contains_total = out[["GUIA", "CLIENTE", "LOCALIDAD", "CIUDAD", "COD RASTREO"]].apply(
        lambda s: s.str.contains(r"\btotal\b", case=False, regex=True, na=False), axis=0
    ).any(axis=1)
    all_text_empty = out[["GUIA", "CLIENTE", "LOCALIDAD", "CIUDAD", "COD RASTREO"]].apply(lambda s: s.str.strip() == "", axis=0).all(axis=1)

    mask_valid = has_any_text & (~contains_total) & (~all_text_empty)
    out = out.loc[mask_valid].reset_index(drop=True)

    total_piezas = int(out["PIEZAS"].sum()) if not out.empty else 0
    out = out[cols_final]
    out = _df_safe_for_excel(out)
    return out, total_piezas


# ======================================================================
#                    Utilidades de formato (openpyxl)
# ======================================================================

def insertar_bloque_firma_ws(ws, total_piezas: Optional[int] = None) -> None:
    """
    Inserta al final de la hoja un bloque de firma y total visualmente alineado como en el formato físico:
    
    Nombre: ___________
    Firma:  ___________                           total: 28
    """
    from openpyxl.styles import Alignment, Font, Border, Side

    # Detecta la última fila con contenido
    max_row = ws.max_row
    max_col = ws.max_column

    # Añade separación vertical para que quede visualmente bien
    start_row = max_row + 3

    # Fuente base
    font_text = Font(name="Segoe UI", size=11)

    # Bloque de firma a la izquierda (compatibilidad con tests y formato histórico)
    ws.cell(row=start_row, column=1, value="Nombre quien recibe:").font = font_text
    ws.cell(row=start_row + 1, column=1, value="Firma quien recibe:").font = font_text

    # Línea de firma con celdas fusionadas y borde inferior
    end_col = max(3, min(max_col, 6))
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=end_col)
    ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=end_col)

    thin = Side(style="thin")
    line_border = Border(bottom=thin)
    for row in (start_row, start_row + 1):
        for col in range(2, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = line_border

    # Total a la derecha (alineado última columna)
    if total_piezas is not None:
        try:
            total_val = int(total_piezas)
        except (TypeError, ValueError):
            total_val = total_piezas
        # Evita escribir en celdas fusionadas (MergedCell) de la fila de firma.
        total_row = start_row + 2
        c_total = ws.cell(row=total_row, column=max_col, value=f"TOTAL: ({total_val})")
        c_total.font = Font(name="Segoe UI", size=11, bold=True)
        c_total.alignment = Alignment(horizontal="right", vertical="center")

def agregar_footer_info_ws(ws, total_piezas: int) -> None:
    """
    Configura el pie de página con el total de piezas y timestamp de generación.
    Se replica en páginas pares para mantener consistencia al imprimir.
    """
    try:
        total_value = int(total_piezas)
    except (TypeError, ValueError):
        total_value = total_piezas

    if isinstance(total_value, (int, float)):
        total_label = f"TOTAL: ({int(total_value)})"
    else:
        total_label = f"TOTAL: ({total_value})"

    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    try:
        ws.oddFooter.left.text = total_label
        ws.oddFooter.right.text = timestamp
        ws.evenFooter.left.text = total_label
        ws.evenFooter.right.text = timestamp
    except Exception:
        from openpyxl.styles import Alignment, Font

        row = ws.max_row + 2
        ws.cell(row=row, column=1, value=total_label).font = Font(name="Segoe UI", size=10, bold=True)
        last_col = ws.max_column or 1
        ts_cell = ws.cell(row=row, column=last_col, value=timestamp)
        ts_cell.font = Font(name="Segoe UI", size=10)
        ts_cell.alignment = Alignment(horizontal="right", vertical="center")
 

def formatear_tabla_ws(ws) -> None:
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    font_base = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = 2
    max_col = ws.max_column
    max_row = ws.max_row

    headers = []
    for c in range(1, max_col + 1):
        headers.append(str(ws.cell(row=header_row, column=c).value or "").strip())

    is_fedex = headers[:6] == ["Tracking Number", "Fecha", "Referencia", "Ciudad", "Receptor", "BULTOS"]
    is_urbano = headers[:6] == ["GUIA", "CLIENTE", "LOCALIDAD", "CIUDAD", "PIEZAS", "COD RASTREO"]

    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    last_is_qty = False
    if headers:
        last_is_qty = headers[-1].upper() in {"BULTOS", "PIEZAS"}

    for r in range(header_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_base
            cell.border = border_all
            if last_is_qty and c == max_col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    if is_fedex:
        min_widths = [22, 12, 18, 18, 22, 8]
    elif is_urbano:
        min_widths = [18, 22, 18, 18, 8, 22]
    else:
        min_widths = [16] * max_col

    for i, w in enumerate(min_widths, start=1):
        if i > max_col:
            break
        col_letter = get_column_letter(i)
        cur = ws.column_dimensions[col_letter].width
        if cur is None or cur < w:
            ws.column_dimensions[col_letter].width = w

    # Limita anchos máximos para evitar corte horizontal en 2 hojas.
    if is_urbano:
        max_widths = [16, 24, 16, 16, 8, 14]
    elif is_fedex:
        max_widths = [18, 12, 16, 16, 20, 8]
    else:
        max_widths = [28] * max_col

    for i, wmax in enumerate(max_widths, start=1):
        if i > max_col:
            break
        col_letter = get_column_letter(i)
        cur = ws.column_dimensions[col_letter].width
        if cur is not None and cur > wmax:
            ws.column_dimensions[col_letter].width = wmax

    # Refuerza impresión a 1 página de ancho.
    try:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        if hasattr(ws, "sheet_properties") and hasattr(ws.sheet_properties, "pageSetUpPr"):
            ws.sheet_properties.pageSetUpPr.fitToPage = True  # type: ignore[attr-defined]
    except Exception:
        pass
