import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from pathlib import Path
from datetime import datetime
import json
import shutil

try:
    import win32com.client
    import win32print
except ImportError:
    win32com = None
    win32print = None

from openpyxl import load_workbook

CONFIG_PATH = Path("config.json")

def cargar_configuracion():
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_configuracion(config):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4)

def obtener_ruta_excel():
    config = cargar_configuracion()
    ruta = config.get("ruta_excel", "")
    if ruta and Path(ruta).exists():
        return ruta
    nueva_ruta = filedialog.askopenfilename(
        title="Selecciona el archivo 'etiqueta pedido.xlsx'",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if nueva_ruta:
        config["ruta_excel"] = nueva_ruta
        guardar_configuracion(config)
        return nueva_ruta
    else:
        messagebox.showerror("Archivo no seleccionado", "No se seleccionó ningún archivo de etiquetas.")
        exit()

def guardar_impresora_en_config(impresora_nombre):
    config = cargar_configuracion()
    config["impresora_por_defecto"] = impresora_nombre
    guardar_configuracion(config)

def obtener_impresora_guardada():
    config = cargar_configuracion()
    return config.get("impresora_por_defecto", "")

def cargar_clientes(path_excel):
    xls = pd.ExcelFile(path_excel)
    df_clientes = xls.parse("Clientes")
    return df_clientes

def buscar_cliente_por_rut(df_clientes, rut):
    fila = df_clientes[df_clientes['rut'] == rut.strip()]
    if not fila.empty:
        datos = fila.iloc[0]
        return {
            "razsoc": datos.get("razsoc", ""),
            "dir": datos.get("dir", ""),
            "comuna": datos.get("comuna", ""),
            "ciudad": datos.get("ciudad", "")
        }
    return None

def imprimir_etiqueta_excel(data, impresora):
    plantilla = Path("plantilla_etiqueta.xlsx")
    if not plantilla.exists():
        messagebox.showerror("Error", "No se encuentra el archivo plantilla_etiqueta.xlsx")
        return

    output_dir = Path("output/etiquetas_excel")
    output_dir.mkdir(parents=True, exist_ok=True)
    salida = output_dir / f"etiqueta_{data['rut']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    shutil.copy(plantilla, salida)

    wb = load_workbook(salida)
    ws = wb.active

    reemplazos = {
        "{{guia}}": data.get("guia", ""),
        "{{rut}}": data.get("rut", ""),
        "{{razsoc}}": data.get("razsoc", ""),
        "{{dir}}": data.get("dir", ""),
        "{{comuna}}": data.get("comuna", ""),
        "{{ciudad}}": data.get("ciudad", ""),
        "{{bultos}}": data.get("bultos", ""),
        "{{transporte}}": data.get("transporte", "")
    }

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for key, val in reemplazos.items():
                    if key in cell.value:
                        cell.value = cell.value.replace(key, val)

    wb.save(salida)

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    ruta_completa = str(salida.resolve())

    try:
        wb_com = excel.Workbooks.Open(ruta_completa)
        wb_com.PrintOut()
        wb_com.Close(SaveChanges=False)
        excel.Quit()
        messagebox.showinfo("Impresión enviada", f"Etiqueta enviada a: {impresora}")
    except Exception as e:
        excel.Quit()
        messagebox.showerror("Error en impresión", str(e))

def crear_editor_etiqueta(df_clientes):
    root = tk.Tk()
    root.title("Impresión de Etiquetas Zebra (Excel)")

    frame = ttk.Frame(root, padding=20)
    frame.grid(row=0, column=0)

    campos = {
        "rut": "RUT",
        "razsoc": "Cliente",
        "dir": "Dirección",
        "comuna": "Comuna",
        "ciudad": "Ciudad",
        "guia": "Guía",
        "bultos": "Bultos",
        "transporte": "Transporte"
    }

    entradas = {}

    for idx, (key, label) in enumerate(campos.items()):
        ttk.Label(frame, text=label + ":").grid(row=idx, column=0, sticky="e", pady=5)
        entry = ttk.Entry(frame, width=40)
        entry.grid(row=idx, column=1, pady=5)
        entradas[key] = entry

    ttk.Label(frame, text="Impresora:").grid(row=len(campos), column=0, sticky="e", pady=5)
    impresoras = [printer['pPrinterName'] for printer in win32print.EnumPrinters(
        win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS, None, 5
    )] if win32print else []

    impresora_var = tk.StringVar()
    impresora_combo = ttk.Combobox(frame, textvariable=impresora_var, values=impresoras, width=38)
    impresora_combo.grid(row=len(campos), column=1, pady=5)

    impresora_guardada = obtener_impresora_guardada()
    if impresora_guardada in impresoras:
        impresora_var.set(impresora_guardada)
        impresora_combo.current(impresoras.index(impresora_guardada))
    elif impresoras:
        impresora_var.set(impresoras[0])

    def cargar_datos_cliente(event=None):
        rut = entradas["rut"].get()
        cliente = buscar_cliente_por_rut(df_clientes, rut)
        if cliente:
            entradas["razsoc"].delete(0, tk.END)
            entradas["razsoc"].insert(0, cliente["razsoc"])
            entradas["dir"].delete(0, tk.END)
            entradas["dir"].insert(0, cliente["dir"])
            entradas["comuna"].delete(0, tk.END)
            entradas["comuna"].insert(0, cliente["comuna"])
            entradas["ciudad"].delete(0, tk.END)
            entradas["ciudad"].insert(0, cliente["ciudad"])
        else:
            messagebox.showerror("RUT no encontrado", "No se encontró el cliente para el RUT ingresado.")

    entradas["rut"].bind("<Return>", cargar_datos_cliente)

    def imprimir():
        data = {k: v.get() for k, v in entradas.items()}
        impresora = impresora_var.get().strip()
        if not impresora:
            messagebox.showwarning("Selecciona una impresora", "Debes seleccionar una impresora.")
            return
        guardar_impresora_en_config(impresora)
        imprimir_etiqueta_excel(data, impresora)

    ttk.Button(frame, text="Imprimir Etiqueta", command=imprimir).grid(row=len(campos)+1, column=0, columnspan=2, pady=15)
    root.mainloop()

# --- Inicio ---
if __name__ == "__main__":
    excel_path = obtener_ruta_excel()
    df_clientes = cargar_clientes(excel_path)
    crear_editor_etiqueta(df_clientes)
