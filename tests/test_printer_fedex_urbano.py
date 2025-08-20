# tests/test_printer_fedex_urbano.py
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook


def _fake_generar_excel_temporal(df, titulo, sheet_name="Hoja"):
    """
    Crea un archivo XLSX minimal con el contenido del DF y devuelve su ruta.
    """
    from tempfile import NamedTemporaryFile
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        p = Path(tmp.name)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # vuelca DF
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    wb.save(p)
    return p


def _assert_bloque_firma_in_wb(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert "Nombre quien recibe:" in values
    assert "Firma quien recibe:" in values


def test_print_fedex_inserta_firma_y_llama_impresora(monkeypatch):
    import app.printer.printer_fedex as mod

    # Mock generar_excel_temporal
    monkeypatch.setattr(mod, "generar_excel_temporal", _fake_generar_excel_temporal)

    called = {"count": 0, "last": None}
    def _fake_send(p):
        called["count"] += 1
        called["last"] = Path(p)

    monkeypatch.setattr(mod, "enviar_a_impresora", _fake_send)

    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    mod.print_fedex(file_path=None, config={}, df=df)

    assert called["count"] == 1
    assert called["last"] and called["last"].exists()

    _assert_bloque_firma_in_wb(called["last"])


def test_print_urbano_inserta_firma_y_llama_impresora(monkeypatch):
    import app.printer.printer_urbano as mod

    # Mock generar_excel_temporal
    monkeypatch.setattr(mod, "generar_excel_temporal", _fake_generar_excel_temporal)

    called = {"count": 0, "last": None}
    def _fake_send(p):
        called["count"] += 1
        called["last"] = Path(p)

    monkeypatch.setattr(mod, "enviar_a_impresora", _fake_send)

    df = pd.DataFrame({"X": ["a", "b", "c"], "Y": [10, 20, 30]})
    mod.print_urbano(file_path=None, config={}, df=df)

    assert called["count"] == 1
    assert called["last"] and called["last"].exists()

    _assert_bloque_firma_in_wb(called["last"])
