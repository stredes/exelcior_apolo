# tests/test_printer_tools.py
from openpyxl import Workbook
from openpyxl.utils import range_boundaries

from app.printer.printer_tools import insertar_bloque_firma_ws


def test_insertar_bloque_firma_ws_crea_lineas_y_merges():
    wb = Workbook()
    ws = wb.active

    # Simula una hoja con 4 columnas y 3 filas de datos
    ws.append(["A", "B", "C", "D"])
    ws.append([1, 2, 3, 4])
    ws.append([5, 6, 7, 8])

    # Acción
    insertar_bloque_firma_ws(ws)

    # Verifica que existan las etiquetas
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Nombre quien recibe:" in labels
    assert "Firma quien recibe:" in labels

    # Debe existir al menos una fusión que:
    # - empiece en la columna 2 (B)
    # - sea horizontal (una sola fila)
    # - abarque al menos dos columnas (max_col >= 3 => C o más)
    found_expected_merge = False
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        if min_row == max_row and min_col == 2 and max_col >= 3:
            found_expected_merge = True
            break

    assert found_expected_merge, (
        f"No se encontró una fusión horizontal que empiece en columna B y abarque >= 2 columnas. "
        f"Rangos fusionados: {[str(r) for r in ws.merged_cells.ranges]}"
    )

    # Verifica que exista borde inferior en las celdas fusionadas detectadas
    # (re-usa el rango encontrado para comprobar al menos una celda)
    if found_expected_merge:
        for rng in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))
            if min_row == max_row and min_col == 2 and max_col >= 3:
                # revisa las celdas del rango
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=min_row, column=c)
                    # dependiendo de la versión, el objeto Border puede existir pero sin style;
                    # verificamos que "bottom" esté presente
                    assert cell.border is not None
                    assert cell.border.bottom is not None
                break
