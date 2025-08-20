# tests/test_printer_listados.py
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook, Workbook


def _fake_generar_excel_temporal(df, titulo, sheet_name="Listado"):
    from tempfile import NamedTemporaryFile
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        p = Path(tmp.name)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # vuelca DF
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    wb.save(p)
    return p


def test_print_listados_aplica_footer(monkeypatch):
    import app.printer.printer_listados as mod

    monkeypatch.setattr(mod, "generar_excel_temporal", _fake_generar_excel_temporal)

    captured = {"path": None}
    def _fake_send(p):
        captured["path"] = Path(p)

    monkeypatch.setattr(mod, "enviar_a_impresora", _fake_send)

    df = pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})
    mod.print_listados(file_path=None, config={}, df=df)

    # Debe haberse guardado un excel
    assert captured["path"] and captured["path"].exists()

    wb = load_workbook(captured["path"])
    ws = wb.active

    # Footer: "Filas: 3" a la izquierda
    assert ws.oddFooter.left.text == "Filas: 3"
    assert ws.evenFooter.left.text == "Filas: 3"

    # Derecha: una fecha con formato dd/mm/YYYY HH:MM (no validamos el valor exacto, solo el formato)
    pat = re.compile(r"\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}")
    assert pat.match(ws.oddFooter.right.text or "")
    assert pat.match(ws.evenFooter.right.text or "")
